import os
import crea_database
import sincronizzatore
import config
import traceback

def main():
    """
    Orchestra l'intero processo di pulizia, creazione del database e sincronizzazione
    per garantire un ambiente di esecuzione consistente. Questo è l'entrypoint
    ufficiale per il processo di sincronizzazione.
    """
    db_path = config.PATH_STORICO_DB
    excel_path = config.PATH_ATTIVITA_PROGRAMMATE
    log_path = "sync_conflicts.log"

    print("--- INIZIO PROCESSO DI SINCRONIZZAZIONE ORCHESTRATO ---")

    # 1. Pulizia dell'ambiente
    print("\n[FASE 1: PULIZIA]")
    try:
        if os.path.exists(db_path):
            os.remove(db_path)
            print(f"  - Rimosso database precedente: {db_path}")
        if os.path.exists(excel_path):
            os.remove(excel_path)
            print(f"  - Rimosso file Excel precedente: {excel_path}")
        if os.path.exists(log_path):
            os.remove(log_path)
            print(f"  - Rimosso log dei conflitti precedente: {log_path}")
        print("Pulizia completata.")
    except Exception as e:
        print(f"ERRORE durante la pulizia: {e}")
        return

    # 2. Creazione del Database
    print("\n[FASE 2: CREAZIONE DATABASE]")
    try:
        crea_database.crea_tabella()
        print(f"Database creato correttamente in: {db_path}")
    except Exception as e:
        print(f"ERRORE CRITICO durante la creazione del database: {e}")
        traceback.print_exc()
        return

    # 3. Sincronizzazione dei Dati
    print("\n[FASE 3: SINCRONIZZAZIONE DATI]")
    try:
        # Crea un file Excel fittizio se non esiste, per scopi di test
        if not os.path.exists(excel_path):
            print(f"  - File Excel non trovato. Creazione di un file di test fittizio in: {excel_path}")
            from openpyxl import Workbook
            from sincronizzatore import EXCEL_TO_DB_MAP
            wb = Workbook()
            ws = wb.active
            ws.title = "A1"
            header_row, data_row = 3, 4
            headers = list(EXCEL_TO_DB_MAP.keys())
            for col_idx, header_text in enumerate(headers, 1):
                ws.cell(row=header_row, column=col_idx, value=header_text)
            pdl_col = headers.index('PdL') + 1
            desc_col = headers.index("DESCRIZIONE\nATTIVITA'") + 1
            ws.cell(row=data_row, column=pdl_col, value='PDL-TEST-01')
            ws.cell(row=data_row, column=desc_col, value='Descrizione di test')
            wb.save(excel_path)
            print("  - File di test creato.")

        success, message = sincronizzatore.sincronizza_dati()
        print(f"Risultato sincronizzazione: {message}")
        if not success:
            print("La sincronizzazione ha riportato un errore.")
            return
    except Exception as e:
        print(f"ERRORE CRITICO durante la sincronizzazione: {e}")
        traceback.print_exc()
        return

    print("\n--- PROCESSO ORCHESTRATO TERMINATO CON SUCCESSO ---")

if __name__ == "__main__":
    main()