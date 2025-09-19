import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import datetime
import os
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from collections import defaultdict
import pyperclip

# --- CONFIGURAZIONE GENERALE ---
PATH_DATABASE = r'\\192.168.11.251\Database_Tecnico_SMI\cartella strumentale condivisa\ALLEGRETTI'
NOME_FILE_DATABASE = "Database_Report_Attivita.xlsm"
NOME_FILE_TEMPLATE = "Database_Template.xlsm"
PATH_GIORNALIERA_BASE = r'\\192.168.11.251\Database_Tecnico_SMI\Giornaliere\Giornaliere 2025'
# MODIFICA: Ora punta al nuovo file gestionale unico
PATH_GESTIONALE = r'C:\Users\Coemi\Desktop\SCRIPT\progetto_questionario_attivita\Gestionale_Tecnici.xlsx'
NOME_FOGLIO_RISPOSTE = "Report Attività Giornaliera (Risposte)"

# Nomi Colonne da Google Sheets
COLONNA_TIMESTAMP = 'Informazioni cronologiche'
COLONNA_UTENTE = 'Nome e Cognome'
COLONNA_DESCRIZIONE = '1. Descrizione PdL'
COLONNA_REPORT = '1. Report Attività'
COLONNA_STATO = '1. Stato attività'
COLONNA_DATA_RIFERIMENTO = 'Data Riferimento Attività'


# --- FUNZIONI ---
def autorizza_google():
    print("Autenticazione a Google Sheets in corso...")
    try:
        scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets', "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
        client = gspread.authorize(creds)
        print("✅ Autenticazione riuscita.")
        return client
    except Exception as e:
        print(f"❌ ERRORE CRITICO: Controlla il file 'credentials.json'. Dettagli: {e}")
        return None

def leggi_dati_da_google(client):
    print(f"\nLeggo i dati dal foglio '{NOME_FOGLIO_RISPOSTE}'...")
    try:
        sheet = client.open(NOME_FOGLIO_RISPOSTE).sheet1
        records = sheet.get_all_records()
        df = pd.DataFrame(records)
        if df.empty: return None
        print(f"✅ Letti {len(df)} record.")
        return df
    except Exception as e:
        print(f"❌ ERRORE durante la lettura dei dati: {e}")
        return None

def carica_database_esistente(percorso_completo):
    if os.path.exists(percorso_completo):
        try:
            print(f"📖 Caricamento database esistente...")
            return pd.read_excel(percorso_completo)
        except Exception as e:
            print(f"⚠️ Errore nel caricamento del database: {e}. Creo un nuovo database...")
    return pd.DataFrame(columns=['PdL', 'Descrizione', 'Stato', 'Tecnico', 'Report', 'Data_Compilazione', 'Data_Riferimento'])

def processa_dati(df_grezzo):
    print("\nProcesso i dati grezzi riga per riga...")
    lista_attivita_pulite = []
    
    colonne_base = [COLONNA_TIMESTAMP, COLONNA_UTENTE, COLONNA_DESCRIZIONE, COLONNA_REPORT, COLONNA_STATO]
    for col in colonne_base:
        if col not in df_grezzo.columns:
            print(f"❌ ERRORE CRITICO: La colonna '{col}' non è stata trovata nel foglio Google!")
            return None
            
    for _, riga in df_grezzo.iterrows():
        if pd.isna(riga[COLONNA_DESCRIZIONE]) or str(riga[COLONNA_DESCRIZIONE]).strip() == '': continue
        
        descrizione_completa = str(riga[COLONNA_DESCRIZIONE])
        pdl_match = re.search(r'PdL (\d{6}/[CS]|\d{6})', descrizione_completa)
        pdl = pdl_match.group(1) if pdl_match else ''
        descrizione_pulita = re.sub(r'PdL \d{6}/?[CS]?\s*[-:]?\s*', '', descrizione_completa).strip()
        
        data_riferimento = None
        if COLONNA_DATA_RIFERIMENTO in riga:
            data_riferimento_str = str(riga.get(COLONNA_DATA_RIFERIMENTO, ''))
            if data_riferimento_str.strip():
                data_riferimento = data_riferimento_str
        
        if not data_riferimento:
            timestamp_str = str(riga.get(COLONNA_TIMESTAMP, ''))
            match = re.search(r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{4})', timestamp_str)
            if match:
                data_riferimento = match.group(1).replace('-', '/').replace('.', '/')
            else:
                data_riferimento = datetime.date.today().strftime('%d/%m/%Y')
        
        nuova_riga = {
            'PdL': pdl, 'Descrizione': descrizione_pulita, 'Stato': riga[COLONNA_STATO],
            'Tecnico': riga[COLONNA_UTENTE], 'Report': riga[COLONNA_REPORT],
            'Data_Compilazione': riga[COLONNA_TIMESTAMP],
            'Data_Riferimento': data_riferimento
        }
        lista_attivita_pulite.append(nuova_riga)
        
    if not lista_attivita_pulite: return None
    df_nuovo = pd.DataFrame(lista_attivita_pulite)
    print(f"✅ Dati processati: create {len(df_nuovo)} attività pulite.")
    return df_nuovo

def aggiorna_database(df_esistente, df_nuovo):
    if df_nuovo is None or df_nuovo.empty:
        print("🟡 Nessun nuovo dato da aggiungere.")
        return df_esistente
    print("🔄 Aggiornamento database con nuovi dati...")
    df_combinato = pd.concat([df_esistente, df_nuovo], ignore_index=True)
    colonne_identificative = ['Tecnico', 'Data_Riferimento', 'PdL']
    df_combinato.sort_values('Data_Compilazione', ascending=True, inplace=True)
    df_combinato.drop_duplicates(subset=colonne_identificative, keep='last', inplace=True)
    df_ordinato = df_combinato.sort_values(by=['Data_Riferimento', 'PdL', 'Tecnico'], ascending=[False, True, True])
    print(f"✅ Database aggiornato: {len(df_ordinato)} record totali.")
    return df_ordinato

def salva_database_excel(df, percorso_salvataggio, nome_file_output, nome_file_template):
    if df is None or df.empty:
        print("🟡 Il database è vuoto, nessun file da salvare.")
        return
    percorso_output = os.path.join(percorso_salvataggio, nome_file_output)
    file_da_caricare = ""
    if os.path.exists(percorso_output):
        file_da_caricare = percorso_output
        print(f"\n💾 Aggiorno il database esistente '{percorso_output}'...")
    else:
        file_da_caricare = nome_file_template
        print(f"\n💾 Creo un nuovo database '{percorso_output}' dal template '{nome_file_template}'...")
    try:
        wb = load_workbook(file_da_caricare, keep_vba=True)
        ws = wb['Database_Attivita']
        ws.delete_rows(2, ws.max_row)
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
        font_bianco = Font(color="FFFFFF")
        for cell in ws[1]: cell.font = font_bianco
        if 'TabellaAttivita' in ws.tables:
            del ws.tables['TabellaAttivita']
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
        max_row, max_col = len(df) + 1, len(df.columns)
        if max_row > 1:
            table_range = f"A1:{get_column_letter(max_col)}{max_row}"
            tab = Table(displayName="TabellaAttivita", ref=table_range)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)
        for column in ws.columns:
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
        os.makedirs(percorso_salvataggio, exist_ok=True)
        wb.save(percorso_output)
        print(f"✅ Database salvato con successo. Le macro sono state preservate.")
    except FileNotFoundError:
        print(f"❌ ERRORE CRITICO: File base non trovato '{file_da_caricare}'.")
    except Exception as e:
        print(f"❌ ERRORE durante il salvataggio del file: {e}")

def get_tecnici_abilitati_e_links():
    try:
        # MODIFICA: Legge dal file gestionale unico
        df_contatti = pd.read_excel(PATH_GESTIONALE, sheet_name='Contatti')
        # Filtra solo i ruoli che devono compilare report
        tecnici_abilitati_df = df_contatti[df_contatti['Ruolo'].isin(['Amministratore', 'Tecnico'])]
        tecnici_abilitati = set(tecnici_abilitati_df['Nome Cognome'].dropna().str.strip())
        links = {row['Nome Cognome']: row.get('Link Attività', '') for _, row in df_contatti.iterrows()}
        return tecnici_abilitati, links
    except Exception as e:
        print(f"\n❌ Errore lettura file gestionale: {e}")
        return None, None

def get_attivita_pianificate(giorno_str, mese_str, anno_str, tecnici_abilitati):
    nome_giornaliera = f"Giornaliera {mese_str}-{anno_str}.xlsm"
    path_giornaliera = os.path.join(PATH_GIORNALIERA_BASE, nome_giornaliera)
    attivita_per_tecnico = defaultdict(set)
    try:
        df_giornaliera = pd.read_excel(path_giornaliera, sheet_name=giorno_str, header=None)
        for _, riga in df_giornaliera.iloc[3:45].iterrows():
            tecnico_short, pdl_text, desc_text = str(riga.get(5, '')).strip(), str(riga.get(9, '')).strip(), str(riga.get(6, '')).strip()
            if not tecnico_short or not pdl_text or not desc_text: continue
            tecnico_full_match = None
            for full_name in tecnici_abilitati:
                parts_completo = full_name.lower().split()
                parts_giornaliera = tecnico_short.lower().split()
                match_trovato = False
                if len(parts_giornaliera) == 1 and parts_giornaliera[0] in parts_completo: match_trovato = True
                elif len(parts_giornaliera) == 2 and parts_giornaliera[1].endswith('.'):
                    cognome_g, iniziale_g = parts_giornaliera[0], parts_giornaliera[1].replace('.', '')
                    if cognome_g in parts_completo and any(p.startswith(iniziale_g) for p in parts_completo if p != cognome_g):
                        match_trovato = True
                if match_trovato:
                    tecnico_full_match = full_name
                    break
            if tecnico_full_match:
                lista_pdl = re.findall(r'(\d{6}/[CS]|\d{6})', pdl_text)
                lista_descrizioni = [line.strip() for line in desc_text.splitlines() if line.strip()]
                if len(lista_pdl) != len(lista_descrizioni):
                    print(f"⚠️ ATTENZIONE: Per {tecnico_full_match}, numero di PdL ({len(lista_pdl)}) e descrizioni ({len(lista_descrizioni)}) non corrispondono.")
                for pdl, desc in zip(lista_pdl, lista_descrizioni):
                    attivita_per_tecnico[tecnico_full_match].add((pdl, desc))
        return attivita_per_tecnico
    except FileNotFoundError:
        return {}
    except Exception as e:
        print(f"\n❌ Errore lettura giornaliera: {e}")
        return None

def verifica_completamento_per_attivita(df_db, data_odierna):
    giorno_str, mese_str, anno_str = str(data_odierna.day), data_odierna.strftime('%m'), str(data_odierna.year)
    data_filtraggio = data_odierna.strftime('%d/%m/%Y')
    print("\n==============================================")
    print("  VERIFICA COMPLETAMENTO REPORT PER ATTIVITÀ  ")
    print("==============================================")
    tecnici_abilitati, links = get_tecnici_abilitati_e_links()
    if tecnici_abilitati is None: return
    attivita_pianificate = get_attivita_pianificate(giorno_str, mese_str, anno_str, tecnici_abilitati)
    if attivita_pianificate is None: return
    pdl_compilati = defaultdict(set)
    if df_db is not None and not df_db.empty:
        report_odierni = df_db[df_db['Data_Riferimento'] == data_filtraggio]
        if not report_odierni.empty:
            pdl_compilati_grouped = report_odierni.groupby('Tecnico')['PdL'].apply(set)
            pdl_compilati.update(pdl_compilati_grouped.to_dict())
            
    testo_per_whatsapp = [f"Controllo giornaliera per la data: {data_filtraggio}", ""]
    tecnici_incompleti_trovati = False
    for tecnico, pdl_pianificati_set in sorted(attivita_pianificate.items()):
        pdl_pianificati = {pdl for pdl, desc in pdl_pianificati_set}
        pdl_completati_tecnico = pdl_compilati.get(tecnico, set())
        pdl_mancanti = pdl_pianificati - pdl_completati_tecnico
        
        if pdl_mancanti:
            tecnici_incompleti_trovati = True
            link_tecnico = links.get(tecnico, "Link non trovato")
            testo_per_whatsapp.append(f"--- *{tecnico.upper()}* ---")
            testo_per_whatsapp.append(f"INCOMPLETO ( _{len(pdl_completati_tecnico)} su {len(pdl_pianificati)}_ )")
            for pdl in sorted(list(pdl_mancanti)):
                testo_per_whatsapp.append(f"  - PdL MANCANTE: {pdl}")
            testo_per_whatsapp.append(f"`{link_tecnico}`")
            testo_per_whatsapp.append("")
            
    if tecnici_incompleti_trovati:
        output_finale = "\n".join(testo_per_whatsapp)
        pyperclip.copy(output_finale)
        print("\n--- RIEPILOGO TECNICI INCOMPLETI ---")
        print(output_finale)
        print("====================================================")
        print("✅ Riepilogo copiato negli appunti!")
        print("   Puoi incollarlo dove vuoi (es. WhatsApp).")
        print("====================================================")
    else:
        print("\n🎉 Tutti i tecnici abilitati e pianificati hanno completato il report! 🎉")


# --- PROGRAMMA PRINCIPALE ---
if __name__ == "__main__":
    print("==============================================")
    print("    AGGIORNAMENTO E VERIFICA REPORT ATTIVITA'   ")
    print("==============================================\n")
    percorso_db_completo = os.path.join(PATH_DATABASE, NOME_FILE_DATABASE)
    db_esistente = carica_database_esistente(percorso_db_completo)
    client_google = autorizza_google()
    db_aggiornato = db_esistente.copy()
    if client_google:
        dati_grezzi = leggi_dati_da_google(client_google)
        if dati_grezzi is not None and not dati_grezzi.empty:
            nuovi_dati = processa_dati(dati_grezzi)
            db_aggiornato = aggiorna_database(db_esistente, nuovi_dati)
    salva_database_excel(db_aggiornato, PATH_DATABASE, NOME_FILE_DATABASE, NOME_FILE_TEMPLATE)
    verifica_completamento_per_attivita(db_aggiornato, datetime.date.today())
    print("\n\n✅ Processo terminato.")