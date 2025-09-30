import openpyxl
from datetime import datetime
import sys

# --- CONFIGURAZIONE ---
FILE_PATH = "ATTIVITA_PROGRAMMATE.xlsm"
SHEETS_TO_MODIFY = ["A1", "A2", "A3", "BLENDING", "CTE"]
COLUMN_NAME = "DataUltimaModifica"
HEADER_ROW = 3  # L'intestazione si trova sulla riga 3
DATA_START_ROW = 4 # I dati iniziano dalla riga 4

# Definiamo un timestamp "punto zero" uguale per tutte le righe
TIME_ZERO = "2025-09-30 20:00:00"

def add_timestamp_column_to_xlsm():
    """
    Aggiunge una colonna di timestamp a fogli specifici di un file .xlsm,
    leggendo l'intestazione dalla riga 3 e popolando i dati dalla riga 4.
    Preserva le macro (VBA) durante l'operazione.

    Questo script è pensato per essere eseguito una sola volta per inizializzare
    il file, ma è costruito in modo sicuro per non duplicare la colonna se
    viene eseguito di nuovo.
    """
    try:
        # Carica il workbook con keep_vba=True per preservare le macro
        print(f"Caricamento del file '{FILE_PATH}' in modalità sicura per le macro...")
        workbook = openpyxl.load_workbook(FILE_PATH, read_only=False, keep_vba=True)
        print("File caricato con successo.")

        for sheet_name in SHEETS_TO_MODIFY:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                print(f"\nAnalisi foglio: '{sheet_name}'...")

                # Leggi l'intestazione dalla riga corretta (HEADER_ROW)
                header_cells = sheet[HEADER_ROW]
                header_values = [cell.value for cell in header_cells]

                # Controlla se la colonna esiste già per evitare duplicati
                if COLUMN_NAME in header_values:
                    print(f"La colonna '{COLUMN_NAME}' esiste già nel foglio '{sheet_name}'. Salto.")
                    continue

                # Trova la prima colonna libera sulla riga dell'intestazione
                new_col_idx = 1
                for cell in header_cells:
                    if cell.value is None:
                        break
                    new_col_idx += 1

                # Aggiungi l'intestazione della nuova colonna
                sheet.cell(row=HEADER_ROW, column=new_col_idx, value=COLUMN_NAME)
                print(f"Aggiunta intestazione '{COLUMN_NAME}' nella colonna {new_col_idx} alla riga {HEADER_ROW}.")

                # Popola la nuova colonna con il timestamp "punto zero"
                # Partiamo dalla riga successiva all'intestazione (DATA_START_ROW)
                rows_populated = 0
                for row_idx in range(DATA_START_ROW, sheet.max_row + 1):
                    # Controlla se la riga è veramente vuota iterando su tutte le sue celle.
                    # Aggiunge il timestamp se anche solo una cella contiene dati.
                    is_row_empty = not any(cell.value is not None for cell in sheet[row_idx])

                    if not is_row_empty:
                        sheet.cell(row=row_idx, column=new_col_idx, value=TIME_ZERO)
                        rows_populated += 1

                print(f"Popolate {rows_populated} righe con il timestamp '{TIME_ZERO}' a partire dalla riga {DATA_START_ROW}.")
            else:
                print(f"\nAvviso: Il foglio '{sheet_name}' non è stato trovato nel file. Salto.")

        # Salva le modifiche sul file originale
        print(f"\nSalvataggio delle modifiche su '{FILE_PATH}'...")
        workbook.save(FILE_PATH)
        print("Operazione completata con successo. Il file è stato aggiornato.")

    except FileNotFoundError:
        print(f"ERRORE CRITICO: Il file '{FILE_PATH}' non è stato trovato.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Si è verificato un errore imprevisto: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    add_timestamp_column_to_xlsm()