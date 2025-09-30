import pandas as pd
import sqlite3
import datetime
import os
import sys
import logging
import openpyxl

# --- CONFIGURAZIONE ---
DB_NAME = "schedario.db"
DB_TABLE_NAME = "attivita_programmate"
EXCEL_FILE_NAME = "ATTIVITA_PROGRAMMATE.xlsm"
EXCEL_SHEET_NAME = "A1"
PRIMARY_KEY = "PdL"
TIMESTAMP_COLUMN = "row_last_modified"
LOCK_FILE = "sync.lock"
DATETIME_FORMAT = '%Y-%m-%d %H:%M:%S'

# --- LOGGING ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler("sync_v2.log", mode='w'), logging.StreamHandler()]
)

# --- MAPPATURA & COLONNE DB ---
HEADER_MAP = {
    'FERM': 'FERM', 'MANUT': 'MANUT', 'PS': 'PS', 'AREA ': 'AREA', 'PdL': 'PdL',
    'IMP.': 'IMP', 'DESCRIZIONE\nATTIVITA\'': 'DESCRIZIONE_ATTIVITA', 'LUN': 'LUN',
    'MAR': 'MAR', 'MER': 'MER', 'GIO': 'GIO', 'VEN': 'VEN', 'STATO\nPdL': 'STATO_PdL',
    'ESE': 'ESE', 'SAIT': 'SAIT', 'PONTEROSSO': 'PONTEROSSO',
    'STATO\nATTIVITA\'': 'STATO_ATTIVITA', 'DATA\nCONTROLLO': 'DATA_CONTROLLO',
    'PERSONALE\nIMPIEGATO': 'PERSONALE_IMPIEGATO', 'PO': 'PO', 'AVVISO': 'AVVISO',
    TIMESTAMP_COLUMN: TIMESTAMP_COLUMN
}
REVERSE_HEADER_MAP = {v: k for k, v in HEADER_MAP.items()}
DB_COLUMNS = [col for col in list(HEADER_MAP.values()) if col != PRIMARY_KEY] + ['Storico']

def create_lock():
    if os.path.exists(LOCK_FILE):
        logging.warning("Lock file exists.")
        return False
    with open(LOCK_FILE, 'w') as f: f.write(str(os.getpid()))
    logging.info("Lock file created.")
    return True

def remove_lock():
    if os.path.exists(LOCK_FILE):
        os.remove(LOCK_FILE)
        logging.info("Lock file removed.")

def load_data_from_excel():
    try:
        df = pd.read_excel(EXCEL_FILE_NAME, sheet_name=EXCEL_SHEET_NAME, header=2, engine='openpyxl', dtype=str).where(pd.notnull, None)
        df.rename(columns=HEADER_MAP, inplace=True)
        if TIMESTAMP_COLUMN not in df.columns: df[TIMESTAMP_COLUMN] = pd.NaT
        df[TIMESTAMP_COLUMN] = pd.to_datetime(df[TIMESTAMP_COLUMN], errors='coerce')
        df.dropna(subset=[PRIMARY_KEY], inplace=True)
        df[PRIMARY_KEY] = df[PRIMARY_KEY].astype(str)
        return df[[col for col in HEADER_MAP.values() if col in df.columns]].set_index(PRIMARY_KEY)
    except Exception as e:
        logging.error(f"Error loading from Excel: {e}", exc_info=True)
        return None

def load_data_from_db():
    try:
        with sqlite3.connect(DB_NAME) as conn:
            return pd.read_sql_query(f"SELECT * FROM {DB_TABLE_NAME}", conn, index_col=PRIMARY_KEY, parse_dates=[TIMESTAMP_COLUMN])
    except Exception as e:
        logging.error(f"Error loading from DB: {e}", exc_info=True)
        return None

def sync_data(df_excel, df_db):
    excel_keys, db_keys = set(df_excel.index), set(df_db.index)
    db_inserts, db_updates, excel_inserts, excel_updates = [], [], [], []
    now = datetime.datetime.now()

    for key in excel_keys.union(db_keys):
        is_in_excel, is_in_db = key in excel_keys, key in db_keys

        if is_in_excel and not is_in_db:
            logging.info(f"[{key}] New in Excel -> DB")
            row_data = df_excel.loc[key].to_dict()
            row_data[TIMESTAMP_COLUMN] = now # New item gets current time
            db_inserts.append(row_data)
        elif is_in_db and not is_in_excel:
            logging.info(f"[{key}] New in DB -> Excel")
            row_data = df_db.loc[key].to_dict()
            row_data[TIMESTAMP_COLUMN] = now # New item gets current time
            excel_inserts.append(row_data)
        elif is_in_excel and is_in_db:
            excel_row, db_row = df_excel.loc[key], df_db.loc[key]
            excel_ts = pd.to_datetime(excel_row.get(TIMESTAMP_COLUMN), errors='coerce')
            db_ts = pd.to_datetime(db_row.get(TIMESTAMP_COLUMN), errors='coerce')

            if pd.isna(excel_ts) and not pd.isna(db_ts): excel_ts = db_ts - datetime.timedelta(seconds=1)
            if pd.isna(db_ts) and not pd.isna(excel_ts): db_ts = excel_ts - datetime.timedelta(seconds=1)

            if pd.isna(excel_ts) or pd.isna(db_ts): continue

            if excel_ts.floor('s') > db_ts.floor('s'):
                logging.info(f"[{key}] Excel is newer -> DB")
                row_data = excel_row.to_dict()
                row_data[TIMESTAMP_COLUMN] = excel_ts # CORRECT: Propagate source timestamp
                db_updates.append(row_data)
            elif db_ts.floor('s') > excel_ts.floor('s'):
                logging.info(f"[{key}] DB is newer -> Excel")
                row_data = db_row.to_dict()
                row_data[TIMESTAMP_COLUMN] = db_ts # CORRECT: Propagate source timestamp
                excel_updates.append(row_data)

        if 'row_data' in locals() and row_data: row_data[PRIMARY_KEY] = key

    if db_inserts or db_updates: commit_to_db(db_inserts, db_updates)
    if excel_inserts or excel_updates: commit_to_excel(excel_inserts, excel_updates, [])

def commit_to_db(inserts, updates):
    with sqlite3.connect(DB_NAME) as conn:
        try:
            with conn:
                for op_list, op_type in [(inserts, "INSERT"), (updates, "UPDATE")]:
                    if not op_list: continue
                    for row_data in op_list:
                        pk = row_data[PRIMARY_KEY]
                        # Prepare data for SQL
                        sql_data = {}
                        for key, value in row_data.items():
                            if key in DB_COLUMNS or key == PRIMARY_KEY:
                                if pd.isna(value):
                                    sql_data[key] = None
                                elif isinstance(value, (datetime.datetime, pd.Timestamp)):
                                    sql_data[key] = value.strftime(DATETIME_FORMAT)
                                else:
                                    sql_data[key] = value

                        update_data = {k: v for k, v in sql_data.items() if k != PRIMARY_KEY}
                        if op_type == "INSERT":
                            update_data[PRIMARY_KEY] = pk
                            cols = ', '.join(f'"{k}"' for k in update_data.keys()); placeholders = ', '.join(['?'] * len(update_data))
                            conn.execute(f"INSERT INTO {DB_TABLE_NAME} ({cols}) VALUES ({placeholders})", list(update_data.values()))
                        elif op_type == "UPDATE":
                            set_clause = ', '.join([f'"{k}" = ?' for k in update_data.keys()])
                            conn.execute(f'UPDATE {DB_TABLE_NAME} SET {set_clause} WHERE "{PRIMARY_KEY}" = ?', list(update_data.values()) + [pk])
                    logging.info(f"Executed {op_type} for {len(op_list)} rows in DB.")
        except sqlite3.Error as e: logging.error(f"DB Error: {e}", exc_info=True)

def commit_to_excel(inserts, updates, deletes):
    logging.info(f"Starting Excel write: {len(inserts)} inserts, {len(updates)} updates, {len(deletes)} deletes.")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_NAME, keep_vba=True)
        ws = wb[EXCEL_SHEET_NAME]
        header_row_idx = 3
        headers = {cell.value: cell.column for cell in ws[header_row_idx] if cell.value}

        if REVERSE_HEADER_MAP[TIMESTAMP_COLUMN] not in headers:
            new_col_idx = ws.max_column + 1
            ws.cell(row=header_row_idx, column=new_col_idx, value=REVERSE_HEADER_MAP[TIMESTAMP_COLUMN])
            headers[REVERSE_HEADER_MAP[TIMESTAMP_COLUMN]] = new_col_idx
            logging.info(f"Created '{TIMESTAMP_COLUMN}' column in Excel.")

        pdl_col_name = REVERSE_HEADER_MAP[PRIMARY_KEY]
        pdl_col_idx = headers[pdl_col_name]
        pdl_map = {str(ws.cell(row=i, column=pdl_col_idx).value): i for i in range(header_row_idx + 1, ws.max_row + 1)}

        if deletes:
            rows_to_delete = sorted([pdl_map[key] for key in deletes if key in pdl_map], reverse=True)
            for row_idx in rows_to_delete: ws.delete_rows(row_idx)
            logging.info(f"Processed {len(rows_to_delete)} deletions from Excel.")

        for op_list, op_type in [(updates, "UPDATE"), (inserts, "INSERT")]:
            if not op_list: continue
            for row_data in op_list:
                row_idx = pdl_map.get(str(row_data[PRIMARY_KEY])) if op_type == "UPDATE" else ws.max_row + 1
                if row_idx:
                    for db_col, value in row_data.items():
                        if db_col in REVERSE_HEADER_MAP and REVERSE_HEADER_MAP[db_col] in headers:
                            col_idx = headers[REVERSE_HEADER_MAP[db_col]]
                            if isinstance(value, (datetime.datetime, pd.Timestamp)): value = value.strftime(DATETIME_FORMAT)
                            ws.cell(row=row_idx, column=col_idx, value=value)
            logging.info(f"Processed {len(op_list)} {op_type}s for Excel.")

        wb.save(EXCEL_FILE_NAME)
        logging.info("Successfully saved Excel file.")
    except Exception as e:
        logging.error(f"Critical error during Excel write: {e}", exc_info=True)

def main():
    if not create_lock(): sys.exit(1)
    try:
        logging.info("--- Starting Sync v2.0 ---")
        df_excel = load_data_from_excel()
        df_db = load_data_from_db()
        if df_excel is None or df_db is None: return

        sync_data(df_excel, df_db)

        logging.info("--- Checking for deletions ---")
        df_excel_after, df_db_after = load_data_from_excel(), load_data_from_db()
        excel_keys, db_keys = set(df_excel_after.index), set(df_db_after.index)

        deleted_from_excel = db_keys - excel_keys
        if deleted_from_excel:
            logging.info(f"Found {len(deleted_from_excel)} rows deleted from Excel. Removing from DB.")
            with sqlite3.connect(DB_NAME) as conn:
                for key in deleted_from_excel:
                    conn.execute(f'DELETE FROM {DB_TABLE_NAME} WHERE "{PRIMARY_KEY}" = ?', (key,))

        deleted_from_db = excel_keys - db_keys
        if deleted_from_db:
             logging.info(f"Found {len(deleted_from_db)} rows deleted from DB. Removing from Excel.")
             commit_to_excel([], [], list(deleted_from_db))

        logging.info("--- Sync Finished ---")
    except Exception as e:
        logging.critical(f"Critical error in main process: {e}", exc_info=True)
    finally:
        remove_lock()

if __name__ == "__main__":
    main()