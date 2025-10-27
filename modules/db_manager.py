import sqlite3
import pandas as pd
import os
import json
import streamlit as st

DB_NAME = "schedario.db"

def get_db_connection():
    """Crea e restituisce una connessione al database."""
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def get_shifts_by_type(shift_type: str) -> pd.DataFrame:
    """Carica i turni di un tipo specifico direttamente dal database."""
    conn = get_db_connection()
    try:
        query = "SELECT * FROM turni WHERE Tipo = ? ORDER BY Data DESC"
        df = pd.read_sql_query(query, conn, params=(shift_type,))
        return df
    except (sqlite3.Error, pd.io.sql.DatabaseError) as e:
        print(f"Errore nel caricare i turni per tipo '{shift_type}': {e}")
        return pd.DataFrame()
    finally:
        if conn:
            conn.close()

def process_and_commit_validated_relazioni(validated_df: pd.DataFrame, validator_id: str) -> bool:
    """Processa le relazioni validate, aggiornandole nel DB."""
    import datetime
    conn = get_db_connection()
    try:
        with conn:
            cursor = conn.cursor()
            for _, row in validated_df.iterrows():
                # Ricostruisci la data di intervento dal formato stringa
                try:
                    data_intervento_iso = pd.to_datetime(row['data_intervento'], format='%d/%m/%Y').isoformat()
                except ValueError:
                    data_intervento_iso = row['data_intervento'] # Mantieni il formato se non è valido

                cursor.execute(
                    """
                    UPDATE relazioni SET
                        data_intervento = ?,
                        tecnico_compilatore = ?,
                        partner = ?,
                        ora_inizio = ?,
                        ora_fine = ?,
                        corpo_relazione = ?,
                        stato = ?,
                        id_validatore = ?,
                        timestamp_validazione = ?
                    WHERE id_relazione = ?
                    """,
                    (
                        data_intervento_iso,
                        row.get('tecnico_compilatore'),
                        row.get('partner'),
                        row.get('ora_inizio'),
                        row.get('ora_fine'),
                        row.get('corpo_relazione'),
                        'Validata', # Imposta lo stato a Validata
                        validator_id,
                        datetime.datetime.now().isoformat(),
                        row['id_relazione']
                    )
                )
        return True
    except sqlite3.Error as e:
        print(f"Errore durante il salvataggio delle relazioni validate: {e}")
        return False
    finally:
        if conn:
            conn.close()

def get_unvalidated_relazioni() -> pd.DataFrame:
    """Carica le relazioni in attesa di validazione ('Inviata')."""
    conn = get_db_connection()
    try:
        query = "SELECT * FROM relazioni WHERE stato = 'Inviata' ORDER BY timestamp_invio ASC"
        df = pd.read_sql_query(query, conn)
        return df
    except (sqlite3.Error, pd.io.sql.DatabaseError) as e:
        print(f"Errore nel caricare le relazioni da validare: {e}")
        return pd.DataFrame()
    finally:
        if conn:
            conn.close()

def get_reports_to_validate() -> pd.DataFrame:
    """Carica tutti i report in attesa di validazione dalla tabella dedicata."""
    conn = get_db_connection()
    try:
        query = "SELECT * FROM report_da_validare ORDER BY data_compilazione ASC"
        df = pd.read_sql_query(query, conn)
        return df
    except (sqlite3.Error, pd.io.sql.DatabaseError) as e:
        print(f"Errore nel caricare i report da validare: {e}")
        return pd.DataFrame()
    finally:
        if conn:
            conn.close()

def delete_reports_by_ids(report_ids: list) -> bool:
    """Cancella una lista di report dalla tabella di validazione."""
    if not report_ids:
        return True
    conn = get_db_connection()
    try:
        with conn:
            placeholders = ','.join('?' for _ in report_ids)
            query = f"DELETE FROM report_da_validare WHERE id_report IN ({placeholders})"
            conn.execute(query, report_ids)
        return True
    except sqlite3.Error as e:
        print(f"Errore durante la cancellazione dei report: {e}")
        return False
    finally:
        if conn:
            conn.close()

def process_and_commit_validated_reports(validated_data: list) -> bool:
    """
    Scrive i report validati nel file Excel 'Database_Report_Attivita.xlsm'
    e poi li rimuove dalla tabella di validazione del database.
    """
    import openpyxl

    EXCEL_REPORT_FILE = "Database_Report_Attivita.xlsm"
    SHEET_NAME = "STRUMENTALE"

    if not validated_data:
        return True

    # 1. Scrivi su Excel
    try:
        # Carica il workbook preservando le macro
        workbook = openpyxl.load_workbook(EXCEL_REPORT_FILE, keep_vba=True)
        sheet = workbook[SHEET_NAME]

        # Cerca la tabella 'Strumentale' nel foglio
        table = sheet.tables.get("Strumentale")
        if not table:
            print(f"ERRORE: La tabella 'Strumentale' non è stata trovata nel foglio '{SHEET_NAME}'. Impossibile aggiungere righe.")
            return False

        # Prepara le righe da aggiungere
        num_rows_added = 0
        for report_dict in validated_data:
            # Ordine delle colonne: PdL, Descrizione, Matricola, Tecnico, Stato, Report, Data_Compilazione, Data_Intervento
            row_to_add = [
                report_dict.get('pdl'),
                report_dict.get('descrizione_attivita'),
                report_dict.get('matricola_tecnico'),
                report_dict.get('nome_tecnico'),
                report_dict.get('stato_attivita'),
                report_dict.get('testo_report'),
                report_dict.get('data_compilazione'),
                report_dict.get('data_riferimento_attivita')
            ]
            sheet.append(row_to_add)
            num_rows_added += 1

        # Se sono state aggiunte righe, espandi il range della tabella
        if num_rows_added > 0:
            from openpyxl.utils import range_boundaries, get_column_letter
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            new_max_row = max_row + num_rows_added
            new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
            table.ref = new_ref

        workbook.save(EXCEL_REPORT_FILE)
        print(f"Aggiunte {len(validated_data)} righe al file Excel '{EXCEL_REPORT_FILE}' e aggiornata la tabella 'Strumentale'.")

    except FileNotFoundError:
        print(f"ERRORE: Il file '{EXCEL_REPORT_FILE}' non è stato trovato.")
        return False
    except KeyError:
        print(f"ERRORE: Il foglio '{SHEET_NAME}' non è stato trovato nel file Excel.")
        return False
    except Exception as e:
        print(f"Errore durante la scrittura sul file Excel: {e}")
        return False

    # 2. Se la scrittura su Excel è andata a buon fine, rimuovi i report dal DB
    report_ids_to_delete = [report['id_report'] for report in validated_data]
    if not delete_reports_by_ids(report_ids_to_delete):
        print("ERRORE CRITICO: I report sono stati scritti su Excel ma non è stato possibile rimuoverli dal database.")
        # In questo scenario, l'utente vedrà ancora i report nella UI, che è meglio
        # che perderli del tutto. Potrà ri-validarli.
        return False

    return True

def get_interventions_for_technician(technician_matricola: str, start_date, end_date) -> pd.DataFrame:
    """Recupera tutti gli interventi per una data matricola in un intervallo di tempo."""
    conn = get_db_connection()
    try:
        query = """
        SELECT
          json_extract(value, '$.PdL') AS PdL,
          json_extract(value, '$.Descrizione') AS Descrizione,
          json_extract(value, '$.Stato') AS Stato,
          json_extract(value, '$.Report') AS Report,
          date(json_extract(value, '$.Data_Riferimento_dt')) AS Data_Riferimento_dt,
          json_extract(value, '$.Tecnico') AS Tecnico,
          json_extract(value, '$.Matricola') AS Matricola
        FROM
          attivita_programmate,
          json_each(attivita_programmate.Storico)
        WHERE
          Matricola = ? AND
          date(Data_Riferimento_dt) BETWEEN date(?) AND date(?)
        ORDER BY
          Data_Riferimento_dt DESC;
        """
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        df = pd.read_sql_query(query, conn, params=(technician_matricola, start_date_str, end_date_str))
        return df
    except (sqlite3.Error, pd.io.sql.DatabaseError) as e:
        print(f"Errore nel recuperare gli interventi per matricola {technician_matricola}: {e}")
        return pd.DataFrame()
    finally:
        if conn:
            conn.close()

def get_technician_performance_data(start_date, end_date) -> pd.DataFrame:
    """Calcola le metriche di performance per i tecnici interrogando direttamente il DB e raggruppando per Matricola."""
    conn = get_db_connection()
    try:
        query = """
        SELECT
            Matricola,
            Tecnico,
            COUNT(*) AS "Totale Interventi",
            CAST(SUM(CASE WHEN Stato = 'TERMINATA' THEN 1 ELSE 0 END) * 100.0 / COUNT(*) AS REAL) AS "Tasso Completamento (%)",
            AVG(julianday(Data_Compilazione) - julianday(Data_Riferimento_dt)) AS "Ritardo Medio Compilazione (gg)",
            SUM(CASE WHEN LENGTH(Report) < 20 THEN 1 ELSE 0 END) AS "Report Sbrigativi"
        FROM (
            SELECT
                json_extract(value, '$.Matricola') AS Matricola,
                json_extract(value, '$.Tecnico') AS Tecnico,
                json_extract(value, '$.Stato') AS Stato,
                json_extract(value, '$.Report') AS Report,
                json_extract(value, '$.Data_Riferimento_dt') AS Data_Riferimento_dt,
                json_extract(value, '$.Data_Compilazione') AS Data_Compilazione
            FROM
                attivita_programmate,
                json_each(attivita_programmate.Storico)
        )
        WHERE
            Matricola IS NOT NULL AND
            date(Data_Riferimento_dt) BETWEEN date(?) AND date(?)
        GROUP BY
            Matricola, Tecnico
        ORDER BY
            "Totale Interventi" DESC;
        """
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        df = pd.read_sql_query(query, conn, params=(start_date_str, end_date_str))

        if not df.empty:
            df['Tasso Completamento (%)'] = df['Tasso Completamento (%)'].map('{:.1f}'.format)
            df['Ritardo Medio Compilazione (gg)'] = df['Ritardo Medio Compilazione (gg)'].map('{:.1f}'.format)
            # Usa il Nome Tecnico come indice per la visualizzazione in UI
            return df.set_index('Tecnico')

        return pd.DataFrame(columns=['Matricola', 'Tecnico', 'Totale Interventi', 'Tasso Completamento (%)', 'Ritardo Medio Compilazione (gg)', 'Report Sbrigativi']).set_index('Tecnico')

    except (sqlite3.Error, pd.io.sql.DatabaseError) as e:
        print(f"Errore nel calcolo delle performance dei tecnici: {e}")
        return pd.DataFrame(columns=['Matricola', 'Tecnico', 'Totale Interventi', 'Tasso Completamento (%)', 'Ritardo Medio Compilazione (gg)', 'Report Sbrigativi']).set_index('Tecnico')
    finally:
        if conn:
            conn.close()

def salva_relazione(dati_relazione: dict) -> bool:
    """Salva una nuova relazione nel database."""
    conn = get_db_connection()
    try:
        with conn:
            cursor = conn.cursor()
            cols = ', '.join(f'"{k}"' for k in dati_relazione.keys())
            placeholders = ', '.join('?' for _ in dati_relazione)
            sql = f"INSERT INTO relazioni ({cols}) VALUES ({placeholders})"
            cursor.execute(sql, list(dati_relazione.values()))
        return True
    except sqlite3.Error as e:
        print(f"Errore durante il salvataggio della relazione nel DB: {e}")
        return False
    finally:
        if conn:
            conn.close()